"""Generate a blank tracker spreadsheet for a problem list.

    python tools/make_template.py --list blind75
    python tools/make_template.py --list neetcode150 --out my-tracker.xlsx
    python tools/make_template.py --all

Writes an .xlsx (and a plain .csv alongside it) into `templates/` by default.
Upload the .xlsx to Google Drive and it opens as a Google Sheet with the
dashboard formulas, the confidence dropdown and its colour rules intact.

Only needed if you want to regenerate the templates — they are committed.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("openpyxl is required to build templates: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
DATA_PATH = ROOT / "data" / "problems.json"
DEFAULT_OUT_DIR = ROOT / "templates"

LIST_TITLES = {
    "blind75": "Blind 75",
    "neetcode150": "NeetCode 150",
    "neetcode250": "NeetCode 250",
}

# The tab is called "Tracker" in every template so `sheet.tab` in config.yml
# keeps working if you switch lists.
SHEET_TITLE = "Tracker"

FONT = "Arial"

# Column order. The nagger only requires Problem, Diff, Cold ✓ (date),
# 1wk Review and 3wk Review — the rest are for you, not the bot.
HEADERS = [
    "#", "Pattern", "Problem", "Diff", "Time Budget", "NeetCode",
    "Cold ✓ (date)", "1wk Review", "3wk Review",
    "Key Insight + Pitfall", "Confidence",
]
COL = {name: i + 1 for i, name in enumerate(HEADERS)}
LAST_COL = 12  # column L — a hidden spare, so the banner rows span A:L
HEADER_ROW = 7
FIRST_DATA_ROW = HEADER_ROW + 1

COLUMN_WIDTHS = {
    "A": 4.38, "B": 19.25, "C": 33.25, "D": 7.0, "E": 14.0, "F": 9.63,
    "G": 11.38, "H": 13.0, "I": 13.0, "J": 39.38, "K": 11.38,
    "L": 11.38, "M": 7.63,
}
ROW_HEIGHTS = {1: 27.75, 3: 18.0, 4: 36.0, 7: 31.5}
DATA_ROW_HEIGHT = 30.0

INK_DARK = "FF1F2937"    # dashboard + header banner
MUTED = "FF6B7280"       # labels, notes, time budget
LINK_BLUE = "FF2563EB"
PATTERN_FILL = "FFEEF2FF"

DIFF_COLORS = {"Easy": "FF10B981", "Medium": "FFF59E0B", "Hard": "FFEF4444"}

TIME_BUDGETS = {
    "Easy": "~45min first / ~5min review",
    "Medium": "~60min first / ~8min review",
    "Hard": "~75min first / ~12min review",
}

# label, colour for the big number, number format
DASHBOARD = [
    ("Total problems", "FF475569", None),
    ("Solved (any confidence)", LINK_BLUE, None),
    ("🔴 Shaky", "FFB91C1C", None),
    ("🟡 OK", "FFB45309", None),
    ("🟢 Strong", "FF047857", None),
    ("% Strong", "FF059669", "0%"),
]

# value, font colour, fill — matches the confidence dropdown
CONFIDENCE_RULES = [
    ("Shaky", "FF991B1B", "FFFEE2E2"),
    ("OK", "FF92400E", "FFFEF3C7"),
    ("Strong", "FF065F46", "FFD1FAE5"),
]

METHOD_NOTE = (
    "4-pass method per problem: (1) cold attempt 15-25min  →  (2) watch "
    "NeetCode + read editorial  →  (3) re-implement from scratch  →  "
    "(4) write key insight + pitfall"
)
REPETITION_NOTE = (
    "Spaced repetition: re-solve cold 1 week later, then 3 weeks later. Mark "
    "confidence after each session. Shaky problems get re-reviewed next week "
    "regardless."
)


def load_problems(list_name: str) -> list[dict]:
    with DATA_PATH.open("r", encoding="utf-8") as f:
        data = json.load(f)
    return [p for p in data["problems"] if list_name in p["lists"]]


def build_workbook(list_name: str, problems: list[dict]):
    pretty = LIST_TITLES[list_name]
    count = len(problems)
    last_row = FIRST_DATA_ROW + count - 1
    conf = f"K{FIRST_DATA_ROW}:K{last_row}"

    wb = Workbook()
    # Replace the workbook's default font so every cell that never gets styled
    # explicitly — the hidden spare column, the covered half of a merged
    # banner, anything you type in later — is Arial 10 rather than Calibri 11.
    # This is the only hook that reaches the cells merging hides; assigning
    # `.font` to one of those is silently ignored.
    wb._fonts[0] = Font(name=FONT, size=10)
    ws = wb.active
    ws.title = SHEET_TITLE

    def span(row: int, first: int, last: int):
        ws.merge_cells(start_row=row, start_column=first, end_row=row, end_column=last)
        return ws.cell(row=row, column=first)

    # ---- title -----------------------------------------------------------
    title = span(1, 1, LAST_COL)
    title.value = f"{pretty} — Tracker"
    title.font = Font(name=FONT, size=16, bold=True)

    # ---- dashboard -------------------------------------------------------
    banner = span(2, 1, LAST_COL)
    banner.value = "DASHBOARD"
    banner.font = Font(name=FONT, size=11, bold=True, color="FFFFFFFF")
    banner.fill = PatternFill("solid", fgColor=INK_DARK)
    banner.alignment = Alignment(horizontal="center", vertical="center")

    formulas = [
        f"={count}",
        f"=COUNTA({conf})",
        f'=COUNTIF({conf},"Shaky")',
        f'=COUNTIF({conf},"OK")',
        f'=COUNTIF({conf},"Strong")',
        f'=IF(COUNTA({conf})=0,0,COUNTIF({conf},"Strong")/{count})',
    ]
    for i, ((label, color, fmt), formula) in enumerate(zip(DASHBOARD, formulas)):
        first = 1 + i * 2  # each stat spans two columns
        label_cell = span(3, first, first + 1)
        label_cell.value = label
        label_cell.font = Font(name=FONT, size=9, bold=True, color=MUTED)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")

        value_cell = span(4, first, first + 1)
        value_cell.value = formula
        value_cell.font = Font(name=FONT, size=20, bold=True, color=color)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            value_cell.number_format = fmt

    for row, note in ((5, METHOD_NOTE), (6, REPETITION_NOTE)):
        cell = span(row, 1, LAST_COL)
        cell.value = note
        cell.font = Font(name=FONT, size=10, italic=True, color=MUTED)

    # ---- header row ------------------------------------------------------
    for name in HEADERS:
        cell = ws.cell(row=HEADER_ROW, column=COL[name], value=name)
        cell.font = Font(name=FONT, size=11, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=INK_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ---- problem rows ----------------------------------------------------
    body = Font(name=FONT, size=10)
    for i, p in enumerate(problems):
        row = FIRST_DATA_ROW + i
        ws.row_dimensions[row].height = DATA_ROW_HEIGHT
        difficulty = p["difficulty"]

        index = ws.cell(row=row, column=COL["#"], value=i + 1)
        index.font = body
        index.alignment = Alignment(horizontal="center", vertical="center")

        pattern = ws.cell(row=row, column=COL["Pattern"], value=p["pattern"])
        pattern.font = Font(name=FONT, size=10, bold=True)
        pattern.fill = PatternFill("solid", fgColor=PATTERN_FILL)
        pattern.alignment = Alignment(vertical="center", wrap_text=True)

        problem = ws.cell(row=row, column=COL["Problem"], value=p["name"])
        problem.font = body
        problem.alignment = Alignment(vertical="center", wrap_text=True)

        diff = ws.cell(row=row, column=COL["Diff"], value=difficulty)
        diff.font = Font(name=FONT, size=10, bold=True,
                         color=DIFF_COLORS.get(difficulty, "FF000000"))
        diff.alignment = Alignment(horizontal="center", vertical="center")

        budget = ws.cell(row=row, column=COL["Time Budget"],
                         value=TIME_BUDGETS.get(difficulty, ""))
        budget.font = Font(name=FONT, size=9, color=MUTED)
        budget.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        link = ws.cell(row=row, column=COL["NeetCode"], value="Open →")
        link.hyperlink = f"https://neetcode.io/problems/{p['neetcode_slug']}"
        link.font = Font(name=FONT, size=10, color=LINK_BLUE, underline="single")
        link.alignment = Alignment(horizontal="center", vertical="center")

        for name in ("Cold ✓ (date)", "1wk Review", "3wk Review"):
            cell = ws.cell(row=row, column=COL[name])
            cell.font = body
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "yyyy-mm-dd"

        insight = ws.cell(row=row, column=COL["Key Insight + Pitfall"])
        insight.font = body
        insight.alignment = Alignment(vertical="top", wrap_text=True)

        confidence = ws.cell(row=row, column=COL["Confidence"])
        confidence.font = Font(name=FONT, size=10, bold=True)
        confidence.alignment = Alignment(horizontal="center", vertical="center")

    # ---- dropdown + colour rules ----------------------------------------
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(value for value, _, _ in CONFIDENCE_RULES) + '"',
        allow_blank=True,
        showDropDown=False,  # False means "show the dropdown arrow"
    )
    ws.add_data_validation(dv)
    dv.add(conf)

    for value, color, fill in CONFIDENCE_RULES:
        ws.conditional_formatting.add(conf, CellIsRule(
            operator="equal",
            formula=[f'"{value}"'],
            font=Font(name=FONT, size=10, bold=True, color=color),
            fill=PatternFill(start_color=fill, end_color=fill, fill_type="solid"),
        ))

    # ---- sizing ----------------------------------------------------------
    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width
    ws.column_dimensions[get_column_letter(LAST_COL)].hidden = True
    for row, height in ROW_HEIGHTS.items():
        ws.row_dimensions[row].height = height

    ws.freeze_panes = ws.cell(row=FIRST_DATA_ROW, column=1)
    return wb


def write_csv(path: Path, problems: list[dict]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for i, p in enumerate(problems, start=1):
            writer.writerow([
                i, p["pattern"], p["name"], p["difficulty"],
                TIME_BUDGETS.get(p["difficulty"], ""),
                f"https://neetcode.io/problems/{p['neetcode_slug']}",
                "", "", "", "", "",
            ])


def build(list_name: str, out: Path) -> None:
    problems = load_problems(list_name)
    if not problems:
        sys.exit(f"No problems found for list '{list_name}'.")
    out.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(list_name, problems).save(out)
    csv_path = out.with_suffix(".csv")
    write_csv(csv_path, problems)
    print(f"Wrote {out} and {csv_path} ({len(problems)} problems)")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--list", choices=sorted(LIST_TITLES), default="blind75")
    parser.add_argument("--out", type=Path, help="output .xlsx path")
    parser.add_argument("--all", action="store_true", help="build every list")
    args = parser.parse_args()

    if args.all:
        for name in ("blind75", "neetcode150", "neetcode250"):
            build(name, DEFAULT_OUT_DIR / f"{name}-tracker.xlsx")
        return 0
    build(args.list, args.out or DEFAULT_OUT_DIR / f"{args.list}-tracker.xlsx")
    return 0


if __name__ == "__main__":
    sys.exit(main())
